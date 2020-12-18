# Program that inverts the row and column of the cells in the spreadsheet.
# For example, the value at row 5, column 3 will be at row 3, column 5, and vice-versa.
# This should be done for all cells in the spreadsheet.
# example.xlsx

import openpyxl

# Loads the workbook.
wb = openpyxl.load_workbook(input('Enter the Spreadsheet you want the cells inverted: ') + '.xlsx')
sheet = wb.active
secondWb = openpyxl.Workbook()
secondSheet = secondWb.get_active_sheet()
# Inverting the cells.
print('Inverting the cells in the workbook...')
for x in range(1, sheet.max_row + 1):
    for y in range(1, sheet.max_column + 1):
        secondSheet.cell(row=y, column=x).value = sheet.cell(row=x, column=y).value

# Saves the workbook.
secondWb.save('InvertedCells.xlsx')
print('Done.')